#!/usr/bin/env python3
"""Build an identifier-free, no-overwrite MEEI participant/media manifest.

Every source member is hashed before cohort labels are joined. Only participant
JPG photographs and MP4 videos enter the participant media table. The two
cohort-level TIF montages are retained as hashed supporting images and can never
be decoded by the dynamic endpoint.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import stat
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence


SCHEMA_VERSION = "meei_participant_manifest_v1"
DATASET = "MEEI_Facial_Palsy_Standard_Set"
METADATA_FILENAME = "SCBS Standard set with data.xlsx"
PAPER_FILENAME = (
    "2020_Greene_The Spectrum of Facial Palsy The MEEI Palsy Photo and "
    "Video Standard Set.pdf"
)
PINNED_METADATA_XLSX_SHA256 = (
    "52f60e8fc73d00bdbb0888ee9b2dc592b2172a234de9049480f66f4e28cfbbd6"
)
PINNED_PAPER_PDF_SHA256 = (
    "57e483f2c44b74d75f4fa033f1e5721dc804b6f404cb15863ee90b0c1a23d243"
)
_SHA256 = re.compile(r"^[0-9a-f]{64}$")
_PARTICIPANT_ID = re.compile(r"^grp_[0-9a-f]{64}$")
_ASSET_ID = re.compile(r"^ast_[0-9a-f]{64}$")
_RECORDING_ID = re.compile(r"^rec_[0-9a-f]{64}$")
_PARTICIPANT_FIELDS = {
    "participant_id", "cohort", "severity_stratum", "binary_label",
    "metadata_status", "media_counts",
}
_MEDIA_FIELDS = {
    "asset_id", "participant_id", "media_type", "source_sha256",
    "file_size_bytes", "dynamic_binary_eligible", "recording_id",
}


@dataclass(frozen=True)
class ExpectedInventory:
    participants: int
    normal_participants: int
    flaccid_participants: int
    synkinetic_participants: int
    videos: int
    photos: int
    supporting_images: int = 0
    supporting_files: int = 0


FROZEN_EXPECTED = ExpectedInventory(
    participants=60,
    normal_participants=10,
    flaccid_participants=25,
    synkinetic_participants=25,
    videos=60,
    photos=480,
    supporting_images=2,
    supporting_files=5,
)


@dataclass
class ManifestAudit:
    assets_hashed: int = 0
    dynamic_eligibility_assigned: int = 0
    labels_joined: int = 0
    metadata_rows_joined: int = 0
    _event: int = 0
    last_hash_or_eligibility_event: int = 0
    first_label_event: int = 0

    def _advance(self) -> int:
        self._event += 1
        return self._event

    def hash_asset(self) -> None:
        self.assets_hashed += 1
        self.last_hash_or_eligibility_event = self._advance()

    def assign_dynamic_eligibility(self) -> None:
        self.dynamic_eligibility_assigned += 1
        self.last_hash_or_eligibility_event = self._advance()

    def join_label(self) -> None:
        self.labels_joined += 1
        event = self._advance()
        if self.first_label_event == 0:
            self.first_label_event = event

    def join_metadata(self) -> None:
        self.metadata_rows_joined += 1
        self._advance()

    def public(self) -> dict[str, int]:
        return {
            "assets_hashed": self.assets_hashed,
            "dynamic_eligibility_assigned": self.dynamic_eligibility_assigned,
            "labels_joined": self.labels_joined,
            "metadata_rows_joined": self.metadata_rows_joined,
        }


@dataclass(frozen=True)
class _Asset:
    path: Path
    relative: str
    source_sha256: str
    file_size_bytes: int
    kind: str
    participant_key: str | None


def _sha256_file(path: Path) -> str:
    info = os.lstat(path)
    if stat.S_ISLNK(info.st_mode) or not stat.S_ISREG(info.st_mode):
        raise ValueError("MEEI source members must be regular non-symlink files")
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sha(value: object, field: str) -> str:
    if not isinstance(value, str) or _SHA256.fullmatch(value) is None:
        raise ValueError(f"{field} must be a lowercase SHA-256")
    return value


def _normalize_key(value: object) -> str:
    if not isinstance(value, str):
        raise ValueError("participant key material must be text")
    normalized = re.sub(r"[^a-z0-9]", "", value.lower())
    if not normalized:
        raise ValueError("participant key is empty after normalization")
    return normalized


def _opaque(prefix: str, value: str) -> str:
    return f"{prefix}_{hashlib.sha256(value.encode('utf-8')).hexdigest()}"


def _cohort_and_key(relative: Path) -> tuple[str, str]:
    parts = relative.parts
    if not parts:
        raise ValueError("media member has no cohort")
    if parts[0] == "Normals" and len(parts) >= 3:
        cohort, participant = "normal", parts[1]
    elif parts[0] == "Flaccid" and len(parts) >= 4:
        cohort, participant = "flaccid", parts[2]
    elif parts[0] == "Synkinetic" and len(parts) >= 4:
        cohort, participant = "synkinetic", parts[2]
    else:
        raise ValueError("participant media does not follow the frozen MEEI tree")
    return cohort, _normalize_key(participant)


def _severity(cohort: str, participant_key: str) -> str:
    if cohort == "normal":
        return "normal"
    for token, output in (
        ("nearnormal", "near_normal"),
        ("complete", "complete"),
        ("severe", "severe"),
        ("moderate", "moderate"),
        ("mild", "mild"),
    ):
        if token in participant_key:
            return output
    raise ValueError("participant severity stratum is not recognized")


def _walk_regular_files(root: Path) -> list[Path]:
    if not root.is_dir() or root.is_symlink():
        raise ValueError("MEEI data root must be a real directory")
    output: list[Path] = []
    for current, directories, files in os.walk(root, followlinks=False):
        current_path = Path(current)
        for name in directories:
            if (current_path / name).is_symlink():
                raise ValueError("MEEI directories must not be symlinks")
        for name in files:
            path = current_path / name
            info = os.lstat(path)
            if stat.S_ISLNK(info.st_mode) or not stat.S_ISREG(info.st_mode):
                raise ValueError("MEEI members must be regular non-symlink files")
            output.append(path)
    return sorted(output, key=lambda item: item.relative_to(root).as_posix())


def _classify(relative: Path) -> tuple[str, str | None]:
    suffix = relative.suffix.lower()
    if suffix in {".jpg", ".mp4"}:
        _cohort, key = _cohort_and_key(relative)
        return ("video" if suffix == ".mp4" else "photo"), key
    if suffix == ".tif" and len(relative.parts) == 2 and relative.parts[0] in {
        "Flaccid", "Synkinetic",
    }:
        return "supporting_image", None
    if relative.as_posix() in {METADATA_FILENAME, PAPER_FILENAME}:
        return "supporting_file", None
    if relative.name == ".DS_Store":
        return "supporting_file", None
    raise ValueError("MEEI contains an unregistered source member")


def _member_fingerprint(assets: Sequence[_Asset]) -> str:
    digest = hashlib.sha256()
    for asset in sorted(assets, key=lambda item: item.relative):
        digest.update(
            f"{asset.relative}\0{asset.file_size_bytes}\0{asset.source_sha256}\n".encode(
                "utf-8"
            )
        )
    return digest.hexdigest()


def build_participant_manifest(
    data_root: Path,
    *,
    metadata_participant_keys: Iterable[str],
    metadata_xlsx_sha256: str,
    paper_pdf_sha256: str,
    expected: ExpectedInventory = FROZEN_EXPECTED,
    audit: ManifestAudit,
) -> dict[str, object]:
    """Hash/inventory every source member before joining labels or metadata."""
    root = Path(data_root)
    metadata_sha = _sha(metadata_xlsx_sha256, "metadata XLSX digest")
    paper_sha = _sha(paper_pdf_sha256, "paper PDF digest")
    assets: list[_Asset] = []
    for path in _walk_regular_files(root):
        relative_path = path.relative_to(root)
        kind, participant_key = _classify(relative_path)
        source_sha = _sha256_file(path)
        audit.hash_asset()
        if kind in {"photo", "video"}:
            audit.assign_dynamic_eligibility()
        assets.append(_Asset(
            path=path,
            relative=relative_path.as_posix(),
            source_sha256=source_sha,
            file_size_bytes=int(os.lstat(path).st_size),
            kind=kind,
            participant_key=participant_key,
        ))

    metadata_asset = next(
        (item for item in assets if item.relative == METADATA_FILENAME), None
    )
    paper_asset = next(
        (item for item in assets if item.relative == PAPER_FILENAME), None
    )
    if metadata_asset is not None and metadata_asset.source_sha256 != metadata_sha:
        raise ValueError("metadata XLSX differs from its pinned digest")
    if paper_asset is not None and paper_asset.source_sha256 != paper_sha:
        raise ValueError("paper PDF differs from its pinned digest")

    participant_assets: dict[str, list[_Asset]] = {}
    cohort_by_key: dict[str, str] = {}
    for asset in assets:
        if asset.participant_key is None:
            continue
        cohort, observed_key = _cohort_and_key(Path(asset.relative))
        if observed_key != asset.participant_key:
            raise AssertionError("participant key classification drifted")
        previous = cohort_by_key.setdefault(observed_key, cohort)
        if previous != cohort:
            raise ValueError("one participant key crosses cohorts")
        participant_assets.setdefault(observed_key, []).append(asset)

    metadata_keys = {_normalize_key(value) for value in metadata_participant_keys}
    if not metadata_keys.issubset(participant_assets):
        raise ValueError("spreadsheet participant exists without a source video")

    participants: list[dict[str, object]] = []
    media: list[dict[str, object]] = []
    video_hashes: set[str] = set()
    for key in sorted(participant_assets):
        cohort = cohort_by_key[key]
        rows = participant_assets[key]
        videos = [item for item in rows if item.kind == "video"]
        photos = [item for item in rows if item.kind == "photo"]
        if len(videos) != 1:
            raise ValueError("every MEEI participant must have exactly one video")
        if videos[0].source_sha256 in video_hashes:
            raise ValueError("video content is duplicated across participants")
        video_hashes.add(videos[0].source_sha256)
        participant_id = _opaque("grp", f"MEEI:participant:{key}")
        label = "unaffected" if cohort == "normal" else "affected"
        audit.join_label()
        metadata_status = "present" if key in metadata_keys else "missing"
        if metadata_status == "present":
            audit.join_metadata()
        participants.append({
            "participant_id": participant_id,
            "cohort": cohort,
            "severity_stratum": _severity(cohort, key),
            "binary_label": label,
            "metadata_status": metadata_status,
            "media_counts": {"video": 1, "photo": len(photos)},
        })
        for asset in sorted(rows, key=lambda item: item.relative):
            eligible = asset.kind == "video"
            media.append({
                "asset_id": _opaque("ast", f"MEEI:asset:{asset.relative}"),
                "participant_id": participant_id,
                "media_type": asset.kind,
                "source_sha256": asset.source_sha256,
                "file_size_bytes": asset.file_size_bytes,
                "dynamic_binary_eligible": eligible,
                "recording_id": (
                    f"rec_{asset.source_sha256}" if eligible else None
                ),
            })

    observed = {
        "participants": len(participants),
        "normal_participants": sum(row["cohort"] == "normal" for row in participants),
        "flaccid_participants": sum(row["cohort"] == "flaccid" for row in participants),
        "synkinetic_participants": sum(row["cohort"] == "synkinetic" for row in participants),
        "videos": sum(row["media_type"] == "video" for row in media),
        "photos": sum(row["media_type"] == "photo" for row in media),
        "supporting_images": sum(item.kind == "supporting_image" for item in assets),
        "supporting_files": sum(item.kind == "supporting_file" for item in assets),
    }
    expected_values = {
        name: int(getattr(expected, name)) for name in expected.__dataclass_fields__
    }
    if observed != expected_values:
        raise ValueError("MEEI source inventory differs from the frozen expected counts")

    counts = {
        "participants": observed["participants"],
        "normal_participants": observed["normal_participants"],
        "facial_palsy_participants": (
            observed["flaccid_participants"] + observed["synkinetic_participants"]
        ),
        "flaccid_participants": observed["flaccid_participants"],
        "synkinetic_participants": observed["synkinetic_participants"],
        "media_assets": len(media),
        "videos": observed["videos"],
        "photos": observed["photos"],
        "supporting_images": observed["supporting_images"],
        "supporting_files": observed["supporting_files"],
        "dynamic_binary_eligible_videos": sum(
            bool(row["dynamic_binary_eligible"]) for row in media
        ),
        "metadata_rows_present": sum(
            row["metadata_status"] == "present" for row in participants
        ),
        "metadata_rows_missing": sum(
            row["metadata_status"] == "missing" for row in participants
        ),
    }
    supporting_image_hashes = sorted(
        item.source_sha256 for item in assets if item.kind == "supporting_image"
    )
    source_digest = hashlib.sha256()
    for row in sorted(media, key=lambda item: str(item["asset_id"])):
        source_digest.update(
            f"{row['media_type']}:{row['source_sha256']}\n".encode("ascii")
        )
    manifest: dict[str, object] = {
        "schema_version": SCHEMA_VERSION,
        "dataset": DATASET,
        "claim_unit": "participant",
        "identity_status": "publisher_participant_directory_one_video_each",
        "endpoint": {
            "target": "normal_vs_facial_palsy",
            "candidate_selection_eligible": False,
            "external_scoring_authorized": False,
            "eligible_media_type": "video",
            "photo_decoding_allowed": False,
            "supporting_image_decoding_allowed": False,
            "hb_grade_available_per_participant": False,
            "eface_score_available_per_participant": False,
            "sunnybrook_score_available_per_participant": False,
        },
        "counts": counts,
        "participants": sorted(participants, key=lambda row: str(row["participant_id"])),
        "media": sorted(media, key=lambda row: str(row["asset_id"])),
        "audit": audit.public(),
        "provenance": {
            "metadata_xlsx_sha256": metadata_sha,
            "paper_pdf_sha256": paper_sha,
            "aggregate_member_manifest_sha256": _member_fingerprint(assets),
            "participant_media_collection_sha256": source_digest.hexdigest(),
            "supporting_image_sha256": supporting_image_hashes,
        },
    }
    validate_participant_manifest(manifest, expected=expected)
    return manifest


def _contains_private_location(value: object, key: str | None = None) -> bool:
    if key is not None and any(token in key.lower() for token in (
        "path", "filename", "stem", "public_key",
    )):
        return True
    if isinstance(value, Mapping):
        return any(
            _contains_private_location(child, str(child_key))
            for child_key, child in value.items()
        )
    if isinstance(value, list):
        return any(_contains_private_location(child) for child in value)
    if isinstance(value, str):
        return value.startswith(("/", "~")) or value.lower().endswith(
            (".mp4", ".jpg", ".tif", ".xlsx", ".pdf")
        )
    return False


def validate_participant_manifest(
    manifest: Mapping[str, object],
    *,
    expected: ExpectedInventory = FROZEN_EXPECTED,
) -> None:
    top = {
        "schema_version", "dataset", "claim_unit", "identity_status",
        "endpoint", "counts", "participants", "media", "audit", "provenance",
    }
    if not isinstance(manifest, Mapping) or set(manifest) != top:
        raise ValueError("MEEI participant manifest top-level schema changed")
    if (
        manifest["schema_version"] != SCHEMA_VERSION
        or manifest["dataset"] != DATASET
        or manifest["claim_unit"] != "participant"
        or _contains_private_location(manifest)
    ):
        raise ValueError("MEEI participant manifest identity or privacy contract failed")
    endpoint = manifest["endpoint"]
    if not isinstance(endpoint, Mapping) or endpoint != {
        "target": "normal_vs_facial_palsy",
        "candidate_selection_eligible": False,
        "external_scoring_authorized": False,
        "eligible_media_type": "video",
        "photo_decoding_allowed": False,
        "supporting_image_decoding_allowed": False,
        "hb_grade_available_per_participant": False,
        "eface_score_available_per_participant": False,
        "sunnybrook_score_available_per_participant": False,
    }:
        raise ValueError("MEEI endpoint contract changed")
    participants = manifest["participants"]
    media = manifest["media"]
    if not isinstance(participants, list) or not isinstance(media, list):
        raise ValueError("MEEI participants and media must be arrays")
    if len(participants) != expected.participants:
        raise ValueError("MEEI participant count differs from the frozen inventory")
    participant_ids: set[str] = set()
    label_by_participant: dict[str, str] = {}
    cohorts: dict[str, int] = {"normal": 0, "flaccid": 0, "synkinetic": 0}
    metadata_present = 0
    for row in participants:
        if not isinstance(row, Mapping) or set(row) != _PARTICIPANT_FIELDS:
            raise ValueError("MEEI participant row schema changed")
        participant_id = row["participant_id"]
        if not isinstance(participant_id, str) or _PARTICIPANT_ID.fullmatch(participant_id) is None:
            raise ValueError("MEEI participant ID is not opaque")
        if participant_id in participant_ids:
            raise ValueError("MEEI participant IDs must be unique")
        participant_ids.add(participant_id)
        cohort = row["cohort"]
        if cohort not in cohorts:
            raise ValueError("MEEI participant cohort is invalid")
        cohorts[str(cohort)] += 1
        expected_label = "unaffected" if cohort == "normal" else "affected"
        if row["binary_label"] != expected_label:
            raise ValueError("MEEI binary label differs from publisher cohort")
        if row["metadata_status"] not in {"present", "missing"}:
            raise ValueError("MEEI metadata status is invalid")
        metadata_present += int(row["metadata_status"] == "present")
        label_by_participant[participant_id] = expected_label
    if cohorts != {
        "normal": expected.normal_participants,
        "flaccid": expected.flaccid_participants,
        "synkinetic": expected.synkinetic_participants,
    }:
        raise ValueError("MEEI cohort counts differ from the frozen inventory")

    assets: set[str] = set()
    recordings: set[str] = set()
    media_by_participant = {
        participant_id: {"video": 0, "photo": 0}
        for participant_id in participant_ids
    }
    for row in media:
        if not isinstance(row, Mapping) or set(row) != _MEDIA_FIELDS:
            raise ValueError("MEEI media row schema changed")
        asset_id = row["asset_id"]
        participant_id = row["participant_id"]
        media_type = row["media_type"]
        source_sha = row["source_sha256"]
        if not isinstance(asset_id, str) or _ASSET_ID.fullmatch(asset_id) is None:
            raise ValueError("MEEI asset ID is not opaque")
        if asset_id in assets:
            raise ValueError("MEEI asset IDs must be unique")
        assets.add(asset_id)
        if participant_id not in participant_ids or media_type not in {"photo", "video"}:
            raise ValueError("MEEI media participant/type is invalid")
        _sha(source_sha, "MEEI source digest")
        if isinstance(row["file_size_bytes"], bool) or not isinstance(
            row["file_size_bytes"], int
        ) or row["file_size_bytes"] <= 0:
            raise ValueError("MEEI media size must be a positive integer")
        eligible = media_type == "video"
        if row["dynamic_binary_eligible"] is not eligible:
            raise ValueError("MEEI dynamic eligibility is not media-only")
        recording_id = row["recording_id"]
        if eligible:
            expected_recording = f"rec_{source_sha}"
            if recording_id != expected_recording or _RECORDING_ID.fullmatch(recording_id) is None:
                raise ValueError("MEEI video recording ID is invalid")
            if recording_id in recordings:
                raise ValueError("MEEI video content/recording IDs must be unique")
            recordings.add(recording_id)
        elif recording_id is not None:
            raise ValueError("MEEI photographs cannot have recording IDs")
        media_by_participant[participant_id][str(media_type)] += 1
    if any(value["video"] != 1 for value in media_by_participant.values()):
        raise ValueError("each MEEI participant requires exactly one video")
    for row in participants:
        if row["media_counts"] != media_by_participant[row["participant_id"]]:
            raise ValueError("MEEI participant media counts do not reconcile")

    counts = manifest["counts"]
    expected_counts = {
        "participants": expected.participants,
        "normal_participants": expected.normal_participants,
        "facial_palsy_participants": (
            expected.flaccid_participants + expected.synkinetic_participants
        ),
        "flaccid_participants": expected.flaccid_participants,
        "synkinetic_participants": expected.synkinetic_participants,
        "media_assets": expected.videos + expected.photos,
        "videos": expected.videos,
        "photos": expected.photos,
        "supporting_images": expected.supporting_images,
        "supporting_files": expected.supporting_files,
        "dynamic_binary_eligible_videos": expected.videos,
        "metadata_rows_present": metadata_present,
        "metadata_rows_missing": expected.participants - metadata_present,
    }
    if counts != expected_counts:
        raise ValueError("MEEI manifest counts do not independently reconcile")
    provenance = manifest["provenance"]
    if not isinstance(provenance, Mapping) or set(provenance) != {
        "metadata_xlsx_sha256", "paper_pdf_sha256",
        "aggregate_member_manifest_sha256", "participant_media_collection_sha256",
        "supporting_image_sha256",
    }:
        raise ValueError("MEEI manifest provenance schema changed")
    for key in (
        "metadata_xlsx_sha256", "paper_pdf_sha256",
        "aggregate_member_manifest_sha256", "participant_media_collection_sha256",
    ):
        _sha(provenance[key], key)
    supporting = provenance["supporting_image_sha256"]
    if (
        not isinstance(supporting, list)
        or len(supporting) != expected.supporting_images
        or supporting != sorted(supporting)
        or any(not isinstance(value, str) or _SHA256.fullmatch(value) is None
               for value in supporting)
    ):
        raise ValueError("MEEI supporting-image digests are invalid")


def _nonempty_workbook_rows(
    rows: Sequence[Sequence[object]],
) -> list[tuple[object, ...]]:
    """Drop only fully empty trailing formatting rows, never interior rows."""
    normalized = [tuple(row) for row in rows]
    while normalized and all(value is None for value in normalized[-1]):
        normalized.pop()
    if any(all(value is None for value in row) for row in normalized):
        raise ValueError("MEEI metadata contains an interior blank row")
    return normalized


def load_metadata_participant_keys(path: Path) -> set[str]:
    checked = Path(path)
    if _sha256_file(checked) != PINNED_METADATA_XLSX_SHA256:
        raise ValueError("MEEI metadata XLSX does not match the pinned artifact")
    from openpyxl import load_workbook

    workbook = load_workbook(checked, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise ValueError("MEEI metadata workbook sheet schema changed")
        rows = _nonempty_workbook_rows(
            list(workbook["Sheet1"].iter_rows(values_only=True))
        )
    finally:
        workbook.close()
    expected_header = (
        "Category", "Sub-category", "#", "Gender", "Age",
        "Side of Paralysis", "Cause of Paralysis",
    )
    if len(rows) != 52 or tuple(rows[0]) != expected_header:
        raise ValueError("MEEI metadata workbook row/header schema changed")
    keys: set[str] = set()
    for row in rows[1:]:
        category, subcategory, number = row[:3]
        if category == "Normal":
            key_material = f"normal{int(number)}"
        elif category == "Flaccid":
            key_material = f"{subcategory}flaccid{int(number)}"
        elif category == "Synkinetic":
            key_material = f"synkinetic{subcategory}{int(number)}"
        else:
            raise ValueError("MEEI metadata category is invalid")
        key = _normalize_key(key_material)
        if key in keys:
            raise ValueError("MEEI metadata participant key is duplicated")
        keys.add(key)
    return keys


def write_private_no_overwrite_json(path: Path, payload: Mapping[str, object]) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    encoded = (
        json.dumps(payload, sort_keys=True, indent=2, ensure_ascii=True, allow_nan=False)
        + "\n"
    ).encode("utf-8")
    with target.open("xb") as stream:
        stream.write(encoded)
        stream.flush()
        os.fsync(stream.fileno())
    os.chmod(target, 0o600)


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-root", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    return parser


def main(argv: list[str] | None = None) -> None:
    args = _parser().parse_args(argv)
    root = args.data_root.expanduser().absolute()
    metadata = root / METADATA_FILENAME
    paper = root / PAPER_FILENAME
    if _sha256_file(paper) != PINNED_PAPER_PDF_SHA256:
        raise ValueError("MEEI paper PDF does not match the pinned artifact")
    audit = ManifestAudit()
    manifest = build_participant_manifest(
        root,
        metadata_participant_keys=load_metadata_participant_keys(metadata),
        metadata_xlsx_sha256=PINNED_METADATA_XLSX_SHA256,
        paper_pdf_sha256=PINNED_PAPER_PDF_SHA256,
        expected=FROZEN_EXPECTED,
        audit=audit,
    )
    write_private_no_overwrite_json(args.output, manifest)
    print(json.dumps({
        "schema_version": SCHEMA_VERSION,
        "participants": manifest["counts"]["participants"],
        "videos": manifest["counts"]["videos"],
        "output_created": True,
    }, sort_keys=True))


if __name__ == "__main__":
    main()
